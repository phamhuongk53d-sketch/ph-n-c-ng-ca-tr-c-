import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from streamlit_gsheets import GSheetsConnection
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==================================================
# CẤU HÌNH STREAMLIT
# ==================================================
st.set_page_config(
    page_title="Hệ thống Trực Công Bằng 2026",
    layout="wide"
)

SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/1IQg-gXpWWL14FjpiPNAaNAOpsRlXv6BWnm9_GOSLOEE/edit?usp=sharing"

conn = st.connection("gsheets", type=GSheetsConnection)

# ==================================================
# HÀM TIỆN ÍCH
# ==================================================
def get_vietnamese_weekday(d: pd.Timestamp) -> str:
    weekdays = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]
    return f"{weekdays[d.weekday()]}- {d.strftime('%d/%m')}"

def get_month_name(month_num: int) -> str:
    months = ["", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5", "Tháng 6",
              "Tháng 7", "Tháng 8", "Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12"]
    return months[month_num]

def get_day_type(weekday: int) -> str:
    """Phân loại ngày trong tuần: weekday (T2-T6) hoặc weekend (T7, CN)"""
    if weekday in [0, 1, 2, 3, 4]:  # T2-T6
        return "weekday"
    else:  # T7, CN
        return "weekend"

def create_excel_report(year, staff_hours_summary, monthly_hours_detail, schedule_data):
    """Tạo file Excel báo cáo"""
    output = io.BytesIO()
    wb = Workbook()
    
    # Sheet 1: Tổng hợp giờ trực cả năm
    ws1 = wb.active
    ws1.title = f"Tổng giờ trực {year}"
    
    # Tiêu đề
    ws1.merge_cells('A1:D1')
    title_cell = ws1['A1']
    title_cell.value = f"BÁO CÁO TỔNG GIỜ TRỰC NĂM {year}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Header bảng
    headers = ["STT", "Nhân viên", "Tổng giờ trực", "Trung bình/tháng"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Dữ liệu tổng giờ
    for idx, (staff, hours) in enumerate(staff_hours_summary.items(), 1):
        ws1.cell(row=idx+3, column=1, value=idx)
        ws1.cell(row=idx+3, column=2, value=staff)
        ws1.cell(row=idx+3, column=3, value=hours)
        ws1.cell(row=idx+3, column=4, value=round(hours/12, 1))
    
    # Điều chỉnh độ rộng cột
    for column in ws1.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Chi tiết giờ trực theo tháng
    ws2 = wb.create_sheet(title="Chi tiết theo tháng")
    
    ws2.merge_cells('A1:E1')
    title_cell2 = ws2['A1']
    title_cell2.value = f"CHI TIẾT GIỜ TRỰC THEO THÁNG NĂM {year}"
    title_cell2.font = Font(bold=True, size=14)
    title_cell2.alignment = Alignment(horizontal='center')
    
    # Header cho bảng chi tiết tháng
    month_headers = ["Tháng"] + list(monthly_hours_detail.keys())
    for col, header in enumerate(month_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header if col == 1 else get_month_name(header))
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Dữ liệu từng nhân viên theo tháng
    all_staff = list(staff_hours_summary.keys())
    for row_idx, staff in enumerate(all_staff, 1):
        ws2.cell(row=row_idx+3, column=1, value=staff)
        for col_idx, month in enumerate(monthly_hours_detail.keys(), 2):
            hours = monthly_hours_detail[month].get(staff, 0)
            ws2.cell(row=row_idx+3, column=col_idx, value=hours)
    
    # Thêm dòng tổng cộng
    total_row = len(all_staff) + 4
    ws2.cell(row=total_row, column=1, value="TỔNG CỘNG")
    ws2.cell(row=total_row, column=1).font = Font(bold=True)
    
    for col_idx, month in enumerate(monthly_hours_detail.keys(), 2):
        month_total = sum(monthly_hours_detail[month].values())
        ws2.cell(row=total_row, column=col_idx, value=month_total)
        ws2.cell(row=total_row, column=col_idx).font = Font(bold=True)
    
    # Điều chỉnh độ rộng cột cho sheet 2
    for column in ws2.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 3-14: Lịch trực từng tháng
    for month in range(1, 13):
        month_schedule = schedule_data.get(month)
        if month_schedule and not month_schedule.empty:
            ws_month = wb.create_sheet(title=f"Tháng {month}")
            
            ws_month.merge_cells('A1:C1')
            title_cell = ws_month['A1']
            title_cell.value = f"LỊCH TRỰC {get_month_name(month).upper()} NĂM {year}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Header cho lịch trực
            schedule_headers = ["Ngày", "Ca ngày (8h-16h)", "Ca đêm (16h-8h)"]
            for col, header in enumerate(schedule_headers, 1):
                cell = ws_month.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Dữ liệu lịch trực
            for idx, row in month_schedule.iterrows():
                ws_month.cell(row=idx+4, column=1, value=row['Ngày'])
                ws_month.cell(row=idx+4, column=2, value=row.get('Ca: 8h00 - 16h00', ''))
                ws_month.cell(row=idx+4, column=3, value=row.get('Ca: 16h00 - 8h00', ''))
            
            # Điều chỉnh độ rộng cột
            for column in ws_month.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_month.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

# ==================================================
# ĐỌC DỮ LIỆU TỪ GOOGLE SHEETS
# ==================================================
try:
    df_raw = conn.read(
        spreadsheet=SPREADSHEET_URL,
        worksheet="Data_Log",
        ttl=0
    )
except Exception:
    df_raw = pd.DataFrame(columns=["Ngày", "Ca", "Nhân viên", "Giờ", "Năm", "Tháng"])

if not df_raw.empty:
    df_raw["Ngày"] = pd.to_datetime(
        df_raw["Ngày"],
        dayfirst=True,
        errors="coerce"
    )
    df_raw = df_raw.dropna(subset=["Ngày"])
    # Thêm cột năm và tháng để dễ phân nhóm
    df_raw["Năm"] = df_raw["Ngày"].dt.year
    df_raw["Tháng"] = df_raw["Ngày"].dt.month
else:
    df_raw = pd.DataFrame(columns=["Ngày", "Ca", "Nhân viên", "Giờ", "Năm", "Tháng"])

# ==================================================
# SIDEBAR – CẤU HÌNH
# ==================================================
with st.sidebar:
    st.header("Cấu hình nhân sự")

    staff_input = st.text_area(
        "Danh sách nhân viên",
        "Trung, Ngà, Liên, Linh, Hà, Bình, Huyền, Thảo, Trang, Hương B"
    )
    staff = [s.strip() for s in staff_input.split(",") if s.strip()]

    st.header("📋 Cấu hình phân công đặc biệt")
    
    st.subheader("Nhân viên chỉ trực ca ngày")
    st.write("Chỉ trực từ Thứ 2 đến Thứ 6:")
    weekday_only_staff = st.multiselect(
        "Chọn nhân viên chỉ trực T2-T6",
        staff,
        default=["Trung", "Ngà"]
    )
    
    st.subheader("Cân bằng giờ trực")
    balance_type = st.radio(
        "Chiến lược cân bằng",
        ["Cân bằng theo tháng", "Cân bằng theo cả năm"],
        help="Cân bằng theo tháng: đảm bảo giờ trực mỗi tháng đồng đều. Cân bằng theo năm: tổng giờ cả năm đồng đều"
    )
    
    max_hours_diff = st.slider(
        "Chênh lệch giờ tối đa/tháng", 
        min_value=0, 
        max_value=40, 
        value=16,
        help="Chênh lệch tối đa giờ trực giữa người nhiều nhất và ít nhất trong cùng tháng"
    )

    st.header("Thời gian phân lịch")
    
    # Cho phép chọn năm
    year = st.selectbox("Năm", [2026, 2025, 2027, 2024], index=0)
    
    # Cho phép chọn tháng bắt đầu và kết thúc
    col1, col2 = st.columns(2)
    with col1:
        start_month = st.selectbox("Tháng bắt đầu", range(1, 13), index=0, format_func=get_month_name)
    with col2:
        end_month = st.selectbox("Tháng kết thúc", range(1, 13), index=11, format_func=get_month_name)
    
    # Tự động tính ngày bắt đầu và kết thúc
    start_date = datetime(year, start_month, 1)
    
    # Tính ngày cuối cùng của tháng kết thúc
    if end_month == 12:
        end_date = datetime(year, 12, 31)
    else:
        next_month = datetime(year, end_month + 1, 1)
        end_date = next_month - timedelta(days=1)
    
    # Hiển thị thông tin đã chọn
    st.info(f"Phân công từ: {start_date.strftime('%d/%m/%Y')} đến: {end_date.strftime('%d/%m/%Y')}")
    
    st.header("Tùy chọn xuất dữ liệu")
    show_all_months = st.checkbox("Hiển thị tất cả các tháng", value=True)
    
    st.header("Điều chỉnh nhân sự")
    st.write("Thêm/xóa nhân sự từ ngày:")
    adjust_date = st.date_input("Ngày điều chỉnh", datetime.now().date())
    action = st.radio("Hành động", ["Thêm nhân sự", "Xóa nhân sự"])
    if action == "Thêm nhân sự":
        new_staff = st.text_input("Nhân viên mới")
    else:
        remove_staff = st.selectbox("Chọn nhân viên cần xóa", staff)

# ==================================================
# TÍNH GIỜ LŨY KẾ ĐẾN TRƯỚC NGÀY BẮT ĐẦU
# ==================================================
history_before = df_raw[df_raw["Ngày"].dt.date < start_date.date()]

luy_ke_hours = {
    s: history_before.loc[
        history_before["Nhân viên"] == s, "Giờ"
    ].sum()
    for s in staff
}

st.subheader(f"📊 Tổng giờ lũy kế đến {start_date.date() - timedelta(days=1)}")
st.dataframe(pd.DataFrame([luy_ke_hours]).T.rename(columns={0: "Số giờ"}))

# ==================================================
# HIỂN THỊ BẢNG TỔNG SỐ GIỜ TRỰC CỦA MỖI NGƯỜI/NĂM
# ==================================================
if not df_raw.empty:
    st.subheader(f"📈 BẢNG TỔNG SỐ GIỜ TRỰC NĂM {year}")
    
    # Tính tổng giờ cho mỗi nhân viên trong năm được chọn
    yearly_total_hours = {}
    monthly_hours_detail = {}
    
    for month in range(1, 13):
        month_data = df_raw[
            (df_raw["Năm"] == year) & 
            (df_raw["Tháng"] == month)
        ]
        
        if not month_data.empty:
            month_hours = month_data.groupby("Nhân viên")["Giờ"].sum().to_dict()
            monthly_hours_detail[month] = month_hours
            
            for staff_member, hours in month_hours.items():
                yearly_total_hours[staff_member] = yearly_total_hours.get(staff_member, 0) + hours
        else:
            monthly_hours_detail[month] = {}
    
    # Thêm những nhân viên không có giờ trực (giờ = 0)
    for staff_member in staff:
        if staff_member not in yearly_total_hours:
            yearly_total_hours[staff_member] = 0
    
    # Sắp xếp theo số giờ
    yearly_total_df = pd.DataFrame(
        list(yearly_total_hours.items()),
        columns=["Nhân viên", f"Tổng giờ trực {year}"]
    ).sort_values(f"Tổng giờ trực {year}", ascending=True)
    
    # Hiển thị bảng
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.dataframe(yearly_total_df, use_container_width=True, hide_index=True)
    
    with col2:
        # Hiển thị thống kê
        st.metric(
            label="Tổng số nhân viên",
            value=len(yearly_total_df)
        )
        st.metric(
            label="Tổng giờ trực cả năm",
            value=int(yearly_total_df[f"Tổng giờ trực {year}"].sum())
        )
        st.metric(
            label="Trung bình giờ/người",
            value=f"{yearly_total_df[f'Tổng giờ trực {year}'].mean():.1f}"
        )
    
    # Hiển thị biểu đồ phân bố giờ trực
    st.markdown("**Phân bố giờ trực:**")
    chart_data = yearly_total_df.set_index("Nhân viên")[f"Tổng giờ trực {year}"]
    st.bar_chart(chart_data)

# ==================================================
# THUẬT TOÁN PHÂN CA CẢI TIẾN VỚI CÂN BẰNG THEO THÁNG
# ==================================================
def generate_schedule_balanced(staff_list, start_date, end_date, weekday_only_staff_list, balance_strategy="month"):
    """
    Tạo lịch trực với cân bằng theo tháng
    balance_strategy: "month" (cân bằng theo tháng) hoặc "year" (cân bằng theo năm)
    """
    rows = []
    
    # Khởi tạo số giờ theo tháng
    month_start = start_date.month
    month_end = end_date.month
    current_year = start_date.year
    
    # Tạo dictionary để theo dõi giờ trực theo tháng
    monthly_hours_tracker = {}
    for month in range(month_start, month_end + 1):
        monthly_hours_tracker[month] = {staff: 0 for staff in staff_list}
    
    # Khởi tạo thời gian có sẵn cho mỗi nhân viên
    available_at = {
        s: datetime.combine(start_date - timedelta(days=1), datetime.min.time())
        for s in staff_list
    }
    
    # Tính số giờ lũy kế theo tháng từ lịch sử
    for staff_member in staff_list:
        for month in range(month_start, month_end + 1):
            month_history = history_before[
                (history_before["Nhân viên"] == staff_member) &
                (history_before["Tháng"] == month) &
                (history_before["Năm"] == current_year)
            ]
            monthly_hours_tracker[month][staff_member] = month_history["Giờ"].sum()
    
    curr_date = start_date
    while curr_date <= end_date:
        current_month = curr_date.month
        current_weekday = curr_date.weekday()  # 0=Monday, 6=Sunday
        base = datetime.combine(curr_date, datetime.min.time())
        
        # Xác định loại ngày
        if current_weekday < 5:  # Thứ 2 đến Thứ 6
            day_type = "weekday"
        else:  # Thứ 7, Chủ nhật
            day_type = "weekend"
        
        # ===== CA NGÀY (08–16) =====
        # Phân loại nhân viên theo điều kiện
        if day_type == "weekday":
            # Ngày trong tuần: cả nhân viên thường và nhân viên chỉ trực T2-T6
            day_candidates = [
                s for s in staff_list
                if available_at[s] <= base.replace(hour=8)
            ]
        else:
            # Cuối tuần: chỉ nhân viên thường (không bao gồm nhân viên chỉ trực T2-T6)
            day_candidates = [
                s for s in staff_list
                if available_at[s] <= base.replace(hour=8) 
                and s not in weekday_only_staff_list
            ]
        
        # Sắp xếp theo chiến lược cân bằng
        if balance_strategy == "month":
            # Cân bằng theo tháng: ưu tiên người có ít giờ nhất trong tháng hiện tại
            day_candidates.sort(
                key=lambda s: (
                    0 if (day_type == "weekday" and s in weekday_only_staff_list) else 1,
                    monthly_hours_tracker[current_month].get(s, 0)
                )
            )
        else:
            # Cân bằng theo năm: ưu tiên người có ít giờ nhất tổng cộng
            day_candidates.sort(
                key=lambda s: (
                    0 if (day_type == "weekday" and s in weekday_only_staff_list) else 1,
                    sum(monthly_hours_tracker[m].get(s, 0) for m in monthly_hours_tracker)
                )
            )
        
        # Kiểm tra chênh lệch giờ trong tháng
        def is_acceptable_candidate(candidate, selected_candidates, month_hours):
            """Kiểm tra xem chọn candidate này có làm chênh lệch giờ quá lớn không"""
            if not selected_candidates:
                return True
            
            # Lấy số giờ của candidate
            candidate_hours = month_hours.get(candidate, 0)
            
            # Tính số giờ trung bình của những người đã chọn
            selected_hours = [month_hours.get(s, 0) for s in selected_candidates]
            avg_selected = sum(selected_hours) / len(selected_hours) if selected_hours else 0
            
            # Kiểm tra chênh lệch
            if abs(candidate_hours - avg_selected) > max_hours_diff:
                return False
            return True
        
        # Chọn 2 người cho ca ngày
        selected_day = []
        for s in day_candidates:
            if len(selected_day) >= 2:
                break
            if s not in selected_day:
                # Kiểm tra chênh lệch giờ
                if is_acceptable_candidate(s, selected_day, monthly_hours_tracker[current_month]):
                    selected_day.append(s)
        
        # Nếu không đủ 2 người thỏa mãn chênh lệch, lấy 2 người đầu tiên
        if len(selected_day) < 2 and day_candidates:
            selected_day = day_candidates[:2]
        
        for s in selected_day:
            rows.append({
                "Ngày": curr_date,
                "Ca": "Ca: 8h00 - 16h00",
                "Nhân viên": s,
                "Giờ": 8,
                "Năm": curr_date.year,
                "Tháng": current_month
            })
            # Cập nhật giờ theo tháng
            monthly_hours_tracker[current_month][s] = monthly_hours_tracker[current_month].get(s, 0) + 8
            available_at[s] = base.replace(hour=16) + timedelta(hours=16)
        
        # ===== CA ĐÊM (16–08) =====
        night_candidates = [
            s for s in staff_list
            if s not in weekday_only_staff_list  # Nhân viên chỉ trực T2-T6 không trực đêm
            and available_at[s] <= base.replace(hour=16)
            and s not in selected_day  # Tránh trùng với ca ngày cùng ngày
        ]
        
        # Sắp xếp theo chiến lược cân bằng
        if balance_strategy == "month":
            night_candidates.sort(key=lambda s: monthly_hours_tracker[current_month].get(s, 0))
        else:
            night_candidates.sort(key=lambda s: sum(monthly_hours_tracker[m].get(s, 0) for m in monthly_hours_tracker))
        
        # Chọn 2 người cho ca đêm
        selected_night = []
        for s in night_candidates:
            if len(selected_night) >= 2:
                break
            if s not in selected_night:
                # Kiểm tra chênh lệch giờ
                if is_acceptable_candidate(s, selected_night, monthly_hours_tracker[current_month]):
                    selected_night.append(s)
        
        # Nếu không đủ 2 người thỏa mãn chênh lệch, lấy 2 người đầu tiên
        if len(selected_night) < 2 and night_candidates:
            selected_night = night_candidates[:2]
        
        for s in selected_night:
            rows.append({
                "Ngày": curr_date,
                "Ca": "Ca: 16h00 - 8h00",
                "Nhân viên": s,
                "Giờ": 16,
                "Năm": curr_date.year,
                "Tháng": current_month
            })
            # Cập nhật giờ theo tháng
            monthly_hours_tracker[current_month][s] = monthly_hours_tracker[current_month].get(s, 0) + 16
            available_at[s] = base + timedelta(days=2)
        
        curr_date += timedelta(days=1)
    
    return pd.DataFrame(rows), monthly_hours_tracker

# ==================================================
# XỬ LÝ ĐIỀU CHỈNH NHÂN SỰ
# ==================================================
def handle_staff_adjustment(df_existing, adjust_date, action, staff_list):
    """Xử lý điều chỉnh nhân sự từ một ngày cụ thể"""
    df_adjusted = df_existing.copy()
    
    # Xóa tất cả dữ liệu từ ngày điều chỉnh trở đi
    mask = df_adjusted["Ngày"].dt.date >= adjust_date
    df_to_keep = df_adjusted[~mask].copy()
    
    return df_to_keep

# ==================================================
# TẠO & HIỂN THỊ LỊCH THEO THÁNG
# ==================================================
if st.button("🚀 TẠO & CẬP NHẬT LỊCH TRỰC"):
    # Hiển thị cấu hình đã chọn
    st.info(f"""
    **Cấu hình phân công:**
    - Nhân viên chỉ trực T2-T6: {', '.join(weekday_only_staff) if weekday_only_staff else 'Không có'}
    - Chiến lược cân bằng: {'Theo tháng' if balance_type == 'Cân bằng theo tháng' else 'Theo cả năm'}
    - Chênh lệch giờ tối đa/tháng: {max_hours_diff} giờ
    """)
    
    # Xử lý điều chỉnh nhân sự nếu có
    if 'adjust_date' in locals() and adjust_date >= start_date.date():
        df_raw = handle_staff_adjustment(df_raw, adjust_date, action, staff)
        # Cập nhật staff list nếu cần
        if action == "Thêm nhân sự" and 'new_staff' in locals() and new_staff:
            staff.append(new_staff.strip())
        elif action == "Xóa nhân sự" and 'remove_staff' in locals() and remove_staff in staff:
            staff.remove(remove_staff)
    
    # Tạo lịch mới với cân bằng
    df_new, monthly_hours_tracker = generate_schedule_balanced(
        staff, 
        start_date, 
        end_date, 
        weekday_only_staff,
        "month" if balance_type == "Cân bằng theo tháng" else "year"
    )
    
    # Kết hợp dữ liệu cũ (trước ngày bắt đầu) và mới
    # Loại bỏ các ngày trùng trong khoảng thời gian mới
    mask_old = (df_raw["Ngày"].dt.date >= start_date.date()) & (df_raw["Ngày"].dt.date <= end_date.date())
    df_old_outside_range = df_raw[~mask_old]
    
    df_total = pd.concat([df_old_outside_range, df_new], ignore_index=True)
    df_total = df_total.sort_values("Ngày").reset_index(drop=True)
    
    # Lưu dữ liệu schedule theo tháng để xuất Excel
    schedule_by_month = {}
    
    # ================== HIỂN THỊ THEO THÁNG ==================
    st.subheader(f"🗓️ LỊCH PHÂN CÔNG NĂM {year}")
    
    # Hiển thị tất cả tháng hoặc chỉ tháng được chọn
    if show_all_months:
        display_months = range(1, 13)
    else:
        display_months = range(start_month, end_month + 1)
    
    for month in display_months:
        # Lọc dữ liệu theo tháng
        month_data = df_total[
            (df_total["Năm"] == year) & 
            (df_total["Tháng"] == month)
        ].copy()
        
        if not month_data.empty:
            st.markdown(f"### 📅 LỊCH PHÂN CÔNG {get_month_name(month).upper()} NĂM {year}")
            
            # Chuẩn bị dữ liệu hiển thị
            df_month_view = month_data.copy()
            
            df_group = (
                df_month_view
                .groupby(["Ngày", "Ca"], as_index=False)["Nhân viên"]
                .apply(lambda x: " ".join(x))
            )
            
            df_pivot = (
                df_group
                .pivot(index="Ngày", columns="Ca", values="Nhân viên")
                .reindex(columns=["Ca: 8h00 - 16h00", "Ca: 16h00 - 8h00"])
                .fillna("")
                .reset_index()
                .sort_values("Ngày")
            )
            
            df_pivot["Ngày"] = df_pivot["Ngày"].apply(get_vietnamese_weekday)
            
            # Lưu schedule theo tháng để xuất Excel
            schedule_by_month[month] = df_pivot
            
            # Hiển thị bảng
            st.table(df_pivot)
            
            # Tính tổng giờ mỗi nhân viên trong tháng
            st.markdown(f"**Bảng tổng số giờ làm {get_month_name(month)}:**")
            month_hours = (
                month_data
                .groupby("Nhân viên")["Giờ"]
                .sum()
                .reset_index()
                .sort_values("Giờ")
            )
            
            # Hiển thị dạng bảng với cột STT
            month_hours_with_index = month_hours.copy()
            month_hours_with_index.insert(0, "STT", range(1, len(month_hours_with_index) + 1))
            
            # Định dạng số giờ
            month_hours_with_index["Giờ"] = month_hours_with_index["Giờ"].astype(int)
            
            # Tính chênh lệch
            if len(month_hours) > 1:
                min_hours = month_hours["Giờ"].min()
                max_hours = month_hours["Giờ"].max()
                diff_hours = max_hours - min_hours
                
                # Hiển thị cảnh báo nếu chênh lệch lớn
                if diff_hours > max_hours_diff:
                    st.warning(f"⚠️ Chênh lệch giờ trong tháng: {diff_hours} giờ (vượt quá giới hạn {max_hours_diff} giờ)")
                else:
                    st.success(f"✓ Chênh lệch giờ trong tháng: {diff_hours} giờ (trong giới hạn)")
            
            st.dataframe(month_hours_with_index, hide_index=True)
            
            # Hiển thị thống kê nhanh
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric(f"Tổng giờ tháng {month}", f"{month_hours['Giờ'].sum():.0f}")
            with col2:
                st.metric("Số người trực", len(month_hours))
            with col3:
                avg_hours = month_hours['Giờ'].mean() if len(month_hours) > 0 else 0
                st.metric("Trung bình/người", f"{avg_hours:.1f}")
            
            st.markdown("---")
    
    # ================== HIỂN THỊ PHÂN TÍCH CÂN BẰNG ==================
    st.subheader("⚖️ PHÂN TÍCH CÂN BẰNG GIỜ TRỰC")
    
    # Tạo bảng tổng hợp chênh lệch theo tháng
    balance_analysis = []
    for month in range(start_month, end_month + 1):
        if month in monthly_hours_tracker:
            month_hours = monthly_hours_tracker[month]
            # Lọc chỉ những người có giờ > 0
            active_staff_hours = {k: v for k, v in month_hours.items() if v > 0}
            if active_staff_hours:
                min_h = min(active_staff_hours.values())
                max_h = max(active_staff_hours.values())
                diff = max_h - min_h
                avg_h = sum(active_staff_hours.values()) / len(active_staff_hours)
                balance_analysis.append({
                    "Tháng": get_month_name(month),
                    "Số NV trực": len(active_staff_hours),
                    "Giờ thấp nhất": min_h,
                    "Giờ cao nhất": max_h,
                    "Chênh lệch": diff,
                    "Trung bình": f"{avg_h:.1f}",
                    "Trạng thái": "✅ Tốt" if diff <= max_hours_diff else "⚠️ Cần điều chỉnh"
                })
    
    if balance_analysis:
        balance_df = pd.DataFrame(balance_analysis)
        st.dataframe(balance_df, hide_index=True)
    
    # ================== CẬP NHẬT DỮ LIỆU TỔNG GIỜ NĂM ==================
    # Tính lại tổng giờ cho mỗi nhân viên trong năm
    yearly_total_hours_new = {}
    monthly_hours_detail_new = {}
    
